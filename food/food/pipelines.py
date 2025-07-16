import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


class FoodPipeline:
    def process_item(self, item, spider) -> dict:
        """
        Processes a scraped item by cleaning and structuring its nutrients data.

        Iterates over the nutrients list and tries to extract the quantity and unit
        from the raw quantity string. Stores cleaned nutrient info in a new
        list and assigns it back to the item.

        :param item: The scraped item containing food data.
        :type item: dict
        :param spider: The spider instance that scraped the item.
        :type spider: scrapy.Spider
        :return: The processed item.
        :rtype: dict
        """
        nutrients = item.get("nutrients", [])
        processed_nutrients = []

        for nutrient in nutrients:
            quantity_text = nutrient.get("raw_quantity", "")
            quantity_parts = quantity_text.split()

            if not quantity_parts:
                continue  # Skip empty or invalid

            numeric = quantity_parts[0].replace(",", "").strip()
            if len(quantity_parts) > 1:
                unit = quantity_parts[1]
            else:
                continue

            # Convert to float or int
            try:
                quantity = float(numeric) if "." in numeric else int(numeric)
            except ValueError:
                quantity = None
                spider.logger.warning(
                    f"Failed to convert quantity: {numeric} in {item['url']} ({nutrient.get('name')})"
                )

            # Store cleaned nutrient info
            processed_nutrients.append(
                {
                    "name": f'{nutrient.get("name", "").strip()} ({unit})',
                    "quantity": quantity,
                    # "group": nutrient.get("group", "").strip()
                }
            )

        item["nutrients"] = processed_nutrients

        return item


class XSLXPipeline:
    def __init__(self) -> None:
        """
        Initialize the XLSX pipeline.

        Sets up an openpyxl workbook and sheet, defines the header row, and sets up
        column category groupings.
        """
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

        self.headers = [
            "Име",
            "Описание",
            "Хранителна група",
            "Калории (к)",
            "Протеин (г)",
            "Въглехидрати (г)",
            "Мазнини (г)",
            "Фибри (г)",
            "Нишесте (г)",
            "Захари (г)",
            "Галактоза (г)",
            "Глюкоза (г)",
            "Захароза (г)",
            "Лактоза (г)",
            "Малтоза (г)",
            "Фруктоза (г)",
            "Бетаин (мг)",
            "Витамин A (IU)",
            "Витамин B1 (Тиамин) (мг)",
            "Витамин B2 (Рибофлавин) (мг)",
            "Витамин B3 (Ниацин) (мг)",
            "Витамин B4 (Холин) (мг)",
            "Витамин B5 (Пантотенова киселина) (мг)",
            "Витамин B6 (Пиридоксин) (мг)",
            "Витамин B9 (Фолиева киселина) (мкг)",
            "Витамин B12 (Кобалкамин) (мкг)",
            "Витамин C (мг)",
            "Витамин D (мкг)",
            "Витамин E (мг)",
            "Витамин K1 (мкг)",
            "Витамин K2 (MK04) (мкг)",
            "Аланин (г)",
            "Аргинин (г)",
            "Аспарагинова киселина (г)",
            "Валин (г)",
            "Глицин (г)",
            "Глутамин (г)",
            "Изолевцин (г)",
            "Левцин (г)",
            "Лизин (г)",
            "Метионин (г)",
            "Пролин (г)",
            "Серин (г)",
            "Тирозин (г)",
            "Треонин (г)",
            "Триптофан (г)",
            "Фенилаланин (г)",
            "Хидроксипролин (г)",
            "Хистидин (г)",
            "Цистин (г)",
            "Мазнини (г)",
            "Мононенаситени мазнини (г)",
            "Полиненаситени мазнини (г)",
            "Наситени мазнини (г)",
            "Трансмазнини (г)",
            "Желязо (мг)",
            "Калий (мг)",
            "Калций (мг)",
            "Манезий (мг)",
            "Манан (г)",
            "Мед (мг)",
            "Натрий (мг)",
            "Селен (мкг)",
            "Флуорид (мг)",
            "Фосфор (мг)",
            "Цинк (мг)",
            "Холестерол (мг)",
            "Фитостероли (мг)",
            "Стимастероли (мг)",
            "Кампестероли (мг)",
            "Бета-ситостероли (мг)",
            "Алкохол (г)",
            "Вода (г)",
            "Кофеин (мг)",
            "Теобромин (мг)",
            "Пепел (г)",
        ]

        # Insert an empty row for category headers
        self.sheet.append([""] * len(self.headers))
        self.sheet.append(self.headers)

        # Define column category groupings (1-based index)
        self.category_ranges = {
            "100 грама съдържат": (4, 7),
            "Въглехидрати": (8, 16),
            "Витамини": (17, 31),
            "Аминокиселини": (32, 50),
            "Мазнини": (51, 55),
            "Минерали": (56, 66),
            "Стероли": (67, 71),
            "Други": (72, 76),
        }

        # Merge cells and assign group names
        for category, (start, end) in self.category_ranges.items():
            self.sheet.merge_cells(
                start_row=1, start_column=start, end_row=1, end_column=end
            )
            cell = self.sheet.cell(row=1, column=start)
            cell.value = category
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def process_item(self, item, spider) -> dict:
        """
        Processes a scraped item and appends its data to the Excel sheet.

        This method extracts relevant information from the item, such as name, description,
        food group, and nutrients. It maps the nutrient names to their quantities and constructs
        a row to be appended to the Excel sheet. Hyperlinks are set for the name and food group
        columns. The updated item is then returned.

        :param item: The scraped item containing food data.
        :type item: dict
        :param spider: The spider instance that scraped the item.
        :type spider: scrapy.Spider
        :return: The processed item.
        :rtype: dict
        """
        nutrient_map = {n["name"]: n["quantity"] for n in item.get("nutrients", [])}
        row = [
            item.get("name", ""),
            item.get("description", ""),
            item.get("food_group", ""),
        ]
        for header in self.headers[3:]:
            row.append(nutrient_map.get(header, ""))

        self.sheet.append(row)
        row_index = self.sheet.max_row
        self.set_hyperlink(row_index, 1, item.get("url", ""))  # Column A: Name
        self.set_hyperlink(
            row_index, 3, item.get("food_group_url", "")
        )  # Column C: Food Group

        return item

    @staticmethod
    def index_to_column_letter(index) -> str:
        """
        Convert a 1-based column index to its corresponding Excel column letter.

        :param int index: 1-based index of the column
        :return: The corresponding Excel column letter
        :rtype: str
        """
        result = ""
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def set_hyperlink(self, row, col, url) -> None:
        """
        Sets a hyperlink on the cell at the given row and column

        :param int row: Row index of the cell
        :param int col: Column index of the cell
        :param str url: URL to link to
        """
        cell = self.sheet.cell(row=row, column=col)
        if url:
            cell.hyperlink = url
            cell.font = Font(color="0563C1", underline="single")

    def close_spider(self, spider):
        # Always show first 2 columns and first two rows
        self.sheet.freeze_panes = "C3"

        for column_cells in self.sheet.columns:
            # Skip merged cells and find a real cell to get the column letter
            for cell in column_cells:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                    column_letter = cell.column_letter
                    break
            else:
                continue  # all cells in this column are merged, skip

            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value), default=0
            )
            self.sheet.column_dimensions[column_letter].width = max_length + 2

        for start_column, end_column in self.category_ranges.values():
            start_letter = self.index_to_column_letter(start_column)
            end_letter = self.index_to_column_letter(end_column - 1)
            self.sheet.column_dimensions.group(
                start_letter, end_letter, outline_level=1, hidden=False
            )

        # Gets the last column as number
        num_columns = len(self.headers)
        # Gets the last column letter
        last_column = get_column_letter(num_columns)
        # Auto filter from A2 to the last column row 2
        self.sheet.auto_filter.ref = f"A2:{last_column}2"

        self.wb.save("foods.xlsx")
